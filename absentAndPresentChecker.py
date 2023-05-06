

'''
1st file  data must be 9 digit and look like this :
221902543
232543234
555452332
642443434
2nd file data : 
 555452332 ghdfgg erte fh rrfyewry rdgfye  555452332  fgdtgdfgdg 555452332 rtergfggdt h 555452332 
 123456789  sadffhfstaghsrthhtdh
 
 
use method : 
This code is a Python program that creates a graphical user interface (GUI) using the PySimpleGUI library. The GUI prompts the user to select two input files, one containing a list of all student IDs, and another containing a list of IDs that are not valid (i.e., garbage IDs). The program reads both files, extracts the 9-digit student IDs from the garbage IDs file, and determines which IDs are missing from the all IDs file. The program then creates a new Excel file with two sheets: one for absent students, and one for present students. The absent sheet contains the list of missing student IDs, while the present sheet contains the list of all IDs minus the missing IDs. 

To use this program, the user should follow these steps:

1. Install the necessary Python libraries (openpyxl, PySimpleGUI).
2. Save the program code in a Python file (e.g., "absent_present_students.py").
3. Open a command prompt or terminal window.
4. Navigate to the directory where the Python file is saved.
5. Run the program by typing "python absent_present_students.py" and pressing Enter.
6. The GUI window will appear. Click the "Browse" buttons to select the input files.
7. Click the "Run" button to execute the program.
8. After the program finishes, a message box will appear indicating that the operation is complete. The output Excel file ("AbsentPresentStudents.xlsx") will be saved in the same directory as the Python file.
9. To exit the program, click the "Exit" button or close the GUI window.

'''

import string
import openpyxl
import PySimpleGUI as sg

# Create the GUI layout
layout = [
    [sg.Text("All student IDs file:"), sg.Input(key="all_ids_file"), sg.FileBrowse()],
    [sg.Text("Garbage IDs file:"), sg.Input(key="garbage_ids_file"), sg.FileBrowse()],
    [sg.Button("Run"), sg.Button("Exit")]
]

# Create the window
window = sg.Window("Absent Present Students", layout)

# Event loop to process user input
while True:
    event, values = window.read()
    if event == "Run":
        all_ids_file = values["all_ids_file"]
        garbage_ids_file = values["garbage_ids_file"]
        
        # Open files for reading
        try:
            all_ids_file = open(all_ids_file, "r")
            garbage_ids_file = open(garbage_ids_file, "r")
        except IOError:
            sg.popup_error("Error: Unable to open file.")
            continue
        
        # Read all IDs from all_ids.txt into a list
        all_ids = []
        for line in all_ids_file:
            all_ids.append(line.strip())

        all_ids_file.close()

        # Read all IDs from garbage_ids.txt into a list
        garbage_ids = []
        for line in garbage_ids_file:
            # Extract 9-digit student IDs from the line
            ids_in_line = [id.translate(str.maketrans('', '', string.punctuation)) for id in line.split() if len(id) == 9 and id.isdigit()]
            garbage_ids += ids_in_line

        garbage_ids_file.close()

        # Find the missing IDs
        missing_ids = set(all_ids) - set(garbage_ids)

        # Create a new Excel file
        workbook = openpyxl.Workbook()

        # Create a new sheet for absent students
        absent_sheet = workbook.create_sheet("Absent Students")

        # Write the absent students to the sheet
        absent_sheet.append(["#", "Student ID"])
        for i, student_id in enumerate(sorted(missing_ids), 1):
            absent_sheet.append([i, student_id])

        # Create a new sheet for present students
        present_sheet = workbook.create_sheet("Present Students")

        # Write the present students to the sheet
        present_sheet.append(["#", "Student ID"])
        for i, student_id in enumerate(sorted(set(all_ids) - missing_ids), 1):
            present_sheet.append([i, student_id])

        # Save the Excel file
        workbook.save("AbsentPresentStudents.xlsx")
        
        sg.popup("Operation complete!")
    elif event == "Exit":
        break

# Close the window
window.close()
