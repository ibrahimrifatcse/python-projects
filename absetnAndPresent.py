#Running the program will create an Excel file. You will get present list and absent list in the file
import string
import openpyxl

# Open files for reading
try:
    all_ids_file = open("EcoAll.txt", "r")   #Keep all student IDs in this file.
    garbage_ids_file = open("Echogarbage.txt", "r") #Put everything in this file including all IDs taken from Zoom chat.
except IOError:
    print("Error: Unable to open file.")
    exit()

# Read all IDs from all_ids.txt into a list
all_ids = []
for line in all_ids_file:
    all_ids.append(line.strip())

all_ids_file.close()

# Read all IDs from garbage_ids.txt into a list
garbage_ids = []
for line in garbage_ids_file:
    # Extract 9-digit student IDs from the line
    ids_in_line = [id.translate(str.maketrans('', '', string.punctuation)) for id in line.split() if len(id) == 9 and id.isdigit()]
    garbage_ids += ids_in_line

garbage_ids_file.close()

# Find the missing IDs
missing_ids = set(all_ids) - set(garbage_ids)

# Create a new Excel file
workbook = openpyxl.Workbook()

# Create a new sheet for absent students
absent_sheet = workbook.create_sheet("Absent Students")

# Write the absent students to the sheet
absent_sheet.append(["#", "Student ID"])
for i, student_id in enumerate(sorted(missing_ids), 1):
    absent_sheet.append([i, student_id])

# Create a new sheet for present students
present_sheet = workbook.create_sheet("Present Students")

# Write the present students to the sheet
present_sheet.append(["#", "Student ID"])
for i, student_id in enumerate(sorted(set(all_ids) - missing_ids), 1):
    present_sheet.append([i, student_id])

# Save the Excel file
workbook.save("AbsentPresentStudents.xlsx")
